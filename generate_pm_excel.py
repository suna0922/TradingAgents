#!/usr/bin/env python3
"""Generate Excel from backtest result: one sheet per PM decision period.

Each sheet contains:
  1. PM Decision header (date, type, trigger, signal)
  2. PM Rules table (rule #, rule type, condition)
  3. Daily Trading Log (date, close, action, shares, position, triggered rules, trade PnL)
  4. Trade executions mapped to daily entries
"""

import json
import sys
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Styles ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
RULE_TYPE_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_FILL = PatternFill("solid", fgColor="4472C4")
SECTION_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=9)
ACTION_BUYS = Font(name="Arial", size=9, color="CC0000", bold=True)  # red=buy (A股红涨)
ACTION_SELL = Font(name="Arial", size=9, color="00AA00", bold=True)  # green=sell
ACTION_HOLD = Font(name="Arial", size=9, color="888888")
RULE_TRIGGERED_FILL = PatternFill("solid", fgColor="FFF2CC")  # yellow highlight
TRADE_FILL = PatternFill("solid", fgColor="E2EFDA")  # light green for trade rows
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)

RULE_TYPE_COLORS = {
    "stop_loss": PatternFill("solid", fgColor="FFC7CE"),
    "take_profit": PatternFill("solid", fgColor="C6EFCE"),
    "reduce_position": PatternFill("solid", fgColor="FFEB9C"),
    "rating_reeval": PatternFill("solid", fgColor="BDD7EE"),
    "observation_anchor": PatternFill("solid", fgColor="D6E4F0"),
    "entry_zone": PatternFill("solid", fgColor="E4DFEC"),
}


def parse_rule(rule_str: str) -> tuple[str, str]:
    """Parse '[rule_type] condition' into (rule_type, condition)."""
    m = re.match(r"\[(\w+)\]\s*(.*)", rule_str)
    if m:
        return m.group(1), m.group(2)
    return "unknown", rule_str


def fmt_date(d: str) -> str:
    """Convert '2026-01-05' or '2026-01-05 00:00:00' to '2026-01-05'."""
    return str(d)[:10]


def fmt_money(v: float) -> float:
    return round(v, 2)


def fmt_pct(v: float) -> str:
    return f"{v*100:.2f}%"


def build_excel(result_path: str, output_path: str):
    with open(result_path) as f:
        data = json.load(f)

    decisions = data.get("decisions", [])
    daily_states = data.get("daily_states", [])
    trade_history = data.get("trade_history", [])
    summary = data.get("summary", {})

    # Build trade lookup: exit_date -> list of trades
    trades_by_exit = {}
    for t in trade_history:
        exit_d = fmt_date(t.get("exit_date", ""))
        if exit_d:
            trades_by_exit.setdefault(exit_d, []).append(t)

    wb = Workbook()

    # ── Sheet 0: Summary ────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "回测总览"

    # Title
    ws0.merge_cells("A1:F1")
    ws0["A1"] = "000423 东阿阿胶 — 回测结果总览"
    ws0["A1"].font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    ws0["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws0.row_dimensions[1].height = 30

    ws0.merge_cells("A2:F2")
    ws0["A2"] = f"期间: {summary.get('start_date','')} ~ {summary.get('end_date','')}  |  大师配置: Buffett/Graham/Lynch/段永平/Munger"
    ws0["A2"].font = Font(name="Arial", size=10, color="666666")
    ws0["A2"].alignment = Alignment(horizontal="center")

    # Summary metrics
    row = 4
    metrics = [
        ("总收益率", fmt_pct(summary.get("total_return", 0))),
        ("基准收益率", fmt_pct(summary.get("benchmark_return", 0))),
        ("超额收益", fmt_pct(summary.get("excess_return", 0))),
        ("最大回撤", fmt_pct(summary.get("max_drawdown", 0))),
        ("夏普比率", f"{summary.get('sharpe_ratio', 0):.2f}"),
        ("胜率", fmt_pct(summary.get("win_rate", 0))),
        ("总交易次数", str(summary.get("total_trades", 0))),
        ("总PnL", f"¥{summary.get('total_pnl', 0):,.2f}"),
        ("初始资金", f"¥{summary.get('initial_capital', 0):,.2f}"),
        ("最终市值", f"¥{summary.get('final_value', 0):,.2f}"),
    ]
    ws0.cell(row=row, column=1, value="指标").font = HEADER_FONT
    ws0.cell(row=row, column=1).fill = HEADER_FILL
    ws0.cell(row=row, column=2, value="数值").font = HEADER_FONT
    ws0.cell(row=row, column=2).fill = HEADER_FILL
    ws0.cell(row=row, column=3, value="说明").font = HEADER_FONT
    ws0.cell(row=row, column=3).fill = HEADER_FILL
    ws0.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    row += 1
    for name, val in metrics:
        ws0.cell(row=row, column=1, value=name).font = DATA_FONT
        ws0.cell(row=row, column=1).border = THIN_BORDER
        ws0.cell(row=row, column=2, value=val).font = Font(name="Arial", size=10, bold=True)
        ws0.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # PM Decisions overview
    row += 2
    ws0.cell(row=row, column=1, value="PM 决策时间线").font = SECTION_FONT
    ws0.cell(row=row, column=1).fill = SECTION_FILL
    ws0.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    headers = ["决策日期", "类型", "触发条件", "信号", "规则数", "Sheet"]
    for c, h in enumerate(headers, 1):
        cell = ws0.cell(row=row, column=c, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    for i, d in enumerate(decisions):
        sheet_name = _sheet_name(d["date"], decisions, i)
        vals = [
            d["date"],
            d.get("type", ""),
            d.get("trigger", ""),
            d.get("signal", ""),
            len(d.get("rules", [])),
            sheet_name,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws0.cell(row=row, column=c, value=v)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
        row += 1

    # Column widths
    ws0.column_dimensions["A"].width = 18
    ws0.column_dimensions["B"].width = 16
    ws0.column_dimensions["C"].width = 24
    ws0.column_dimensions["D"].width = 14
    ws0.column_dimensions["E"].width = 10
    ws0.column_dimensions["F"].width = 20

    # ── Sheets 1..N: One per decision period ────────────────────────────
    for i, dec in enumerate(decisions):
        dec_date = fmt_date(dec["date"])
        next_date = fmt_date(decisions[i + 1]["date"]) if i + 1 < len(decisions) else None

        sheet_name = _sheet_name(dec_date, decisions, i)
        ws = wb.create_sheet(title=sheet_name)

        # ── Section 1: PM Decision Header ───────────────────────────────
        ws.merge_cells("A1:H1")
        ws["A1"] = f"PM 决策: {dec_date}" + (f" ~ {next_date}" if next_date else " ~ 结束")
        ws["A1"].font = Font(name="Arial", size=13, bold=True, color="1F4E79")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Decision info
        row = 3
        info_items = [
            ("决策日期", dec_date),
            ("决策类型", dec.get("type", "")),
            ("触发条件", dec.get("trigger", "")),
            ("PM信号", dec.get("signal", "")),
            ("规则数量", str(len(dec.get("rules", [])))),
        ]
        for label, val in info_items:
            ws.cell(row=row, column=1, value=label).font = Font(name="Arial", size=9, bold=True, color="666666")
            ws.cell(row=row, column=2, value=val).font = Font(name="Arial", size=10, bold=True)
            row += 1

        # ── Section 2: PM Rules Table ───────────────────────────────────
        row += 1
        rules = dec.get("rules", [])
        ws.cell(row=row, column=1, value=f"📋 PM 交易规则 ({len(rules)} 条)").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

        rule_headers = ["#", "规则类型", "条件表达式", "说明"]
        for c, h in enumerate(rule_headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
        row += 1

        rule_descriptions = {
            "stop_loss": "止损线：触及即卖出",
            "take_profit": "止盈线：触及即卖出",
            "reduce_position": "减仓信号：部分卖出",
            "rating_reeval": "评级重估：触发后重新分析基本面",
            "observation_anchor": "观察锚点：记录但不交易",
            "entry_zone": "入场区域：买入信号",
        }

        for idx, rule_str in enumerate(rules):
            rtype, condition = parse_rule(rule_str)
            ws.cell(row=row, column=1, value=idx + 1).font = DATA_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

            cell = ws.cell(row=row, column=2, value=rtype)
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.fill = RULE_TYPE_COLORS.get(rtype, RULE_TYPE_FILL)
            cell.border = THIN_BORDER

            ws.cell(row=row, column=3, value=condition).font = Font(name="Consolas", size=9)
            ws.cell(row=row, column=3).border = THIN_BORDER

            desc = rule_descriptions.get(rtype, "")
            ws.cell(row=row, column=4, value=desc).font = Font(name="Arial", size=9, color="666666")
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
            ws.cell(row=row, column=4).border = THIN_BORDER
            row += 1

        # ── Section 3: Daily Trading Log ────────────────────────────────
        row += 2
        ws.cell(row=row, column=1, value=f"📊 每日交易日志 ({dec_date} ~ {next_date or '结束'})").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

        log_headers = [
            "日期", "收盘价", "涨跌幅", "操作", "操作股数", "持仓股数",
            "持仓市值", "总资产", "触发的规则", "交易PnL", "平仓原因"
        ]
        for c, h in enumerate(log_headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[row].height = 20
        row += 1

        # Filter daily states for this period
        prev_close = None
        for ds in daily_states:
            ds_date = fmt_date(ds["date"])
            # Include this day if it's >= dec_date and (no next_date or < next_date)
            if ds_date < dec_date:
                prev_close = ds["close"]
                continue
            if next_date and ds_date >= next_date:
                break

            close = ds.get("close", 0)
            change_pct = ""
            if prev_close and prev_close > 0:
                change_pct = f"{(close - prev_close) / prev_close * 100:+.2f}%"

            action = ds.get("action", "HOLD")
            action_shares = ds.get("action_shares", 0)
            shares = ds.get("shares", 0)
            pos_value = ds.get("position_value", 0)
            total_value = ds.get("total_value", 0)
            triggered = ds.get("triggered_rules", [])

            # Check for trades closing on this day
            day_trades = trades_by_exit.get(ds_date, [])
            trade_pnl = sum(t.get("pnl", 0) for t in day_trades)
            exit_reasons = "; ".join(t.get("exit_reason", "") for t in day_trades if t.get("exit_reason"))

            # Write row
            ws.cell(row=row, column=1, value=ds_date).font = DATA_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER

            ws.cell(row=row, column=2, value=fmt_money(close)).font = DATA_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=2).number_format = "¥#,##0.00"

            ws.cell(row=row, column=3, value=change_pct).font = DATA_FONT
            ws.cell(row=row, column=3).border = THIN_BORDER
            # Color the change: red for up, green for down (A股惯例)
            if change_pct:
                if change_pct.startswith("+"):
                    ws.cell(row=row, column=3).font = Font(name="Arial", size=9, color="CC0000")
                elif change_pct.startswith("-"):
                    ws.cell(row=row, column=3).font = Font(name="Arial", size=9, color="00AA00")

            # Action with color
            action_cell = ws.cell(row=row, column=4, value=action)
            action_cell.border = THIN_BORDER
            action_cell.alignment = Alignment(horizontal="center")
            if action == "BUY":
                action_cell.font = ACTION_BUYS
            elif action == "SELL":
                action_cell.font = ACTION_SELL
            else:
                action_cell.font = ACTION_HOLD

            ws.cell(row=row, column=5, value=action_shares if action_shares else "").font = DATA_FONT
            ws.cell(row=row, column=5).border = THIN_BORDER
            ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")

            ws.cell(row=row, column=6, value=shares).font = DATA_FONT
            ws.cell(row=row, column=6).border = THIN_BORDER
            ws.cell(row=row, column=6).alignment = Alignment(horizontal="right")

            ws.cell(row=row, column=7, value=fmt_money(pos_value)).font = DATA_FONT
            ws.cell(row=row, column=7).border = THIN_BORDER
            ws.cell(row=row, column=7).number_format = "¥#,##0.00"

            ws.cell(row=row, column=8, value=fmt_money(total_value)).font = DATA_FONT
            ws.cell(row=row, column=8).border = THIN_BORDER
            ws.cell(row=row, column=8).number_format = "¥#,##0.00"

            # Triggered rules — highlight if any
            triggered_str = "\n".join(triggered) if triggered else ""
            rule_cell = ws.cell(row=row, column=9, value=triggered_str)
            rule_cell.font = Font(name="Consolas", size=8, color="CC6600") if triggered else DATA_FONT
            rule_cell.border = THIN_BORDER
            rule_cell.alignment = Alignment(wrap_text=True, vertical="top")
            if triggered:
                rule_cell.fill = RULE_TRIGGERED_FILL

            # Trade PnL
            if day_trades:
                ws.cell(row=row, column=10, value=fmt_money(trade_pnl)).font = Font(name="Arial", size=9, bold=True, color="00AA00")
                ws.cell(row=row, column=10).fill = TRADE_FILL
            else:
                ws.cell(row=row, column=10, value="").font = DATA_FONT
            ws.cell(row=row, column=10).border = THIN_BORDER
            ws.cell(row=row, column=10).number_format = "¥#,##0.00"

            # Exit reason
            ws.cell(row=row, column=11, value=exit_reasons).font = Font(name="Consolas", size=8, color="666666")
            ws.cell(row=row, column=11).border = THIN_BORDER
            ws.cell(row=row, column=11).alignment = Alignment(wrap_text=True, vertical="top")
            if day_trades:
                ws.cell(row=row, column=11).fill = TRADE_FILL

            # Highlight entire row if trade happened
            if day_trades:
                for c in range(1, 12):
                    if c not in (9, 10, 11):  # Don't double-fill
                        ws.cell(row=row, column=c).fill = TRADE_FILL

            prev_close = close
            row += 1

        # ── Section 4: Trade Executions in this period ──────────────────
        row += 2
        period_trades = []
        for t in trade_history:
            exit_d = fmt_date(t.get("exit_date", ""))
            if exit_d and exit_d >= dec_date and (not next_date or exit_d < next_date):
                period_trades.append(t)

        ws.cell(row=row, column=1, value=f"💰 本期间交易执行 ({len(period_trades)} 笔)").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

        if period_trades:
            trade_headers = ["入场日期", "平仓日期", "方向", "股数", "入场价", "平仓价", "PnL", "收益率", "平仓原因"]
            for c, h in enumerate(trade_headers, 1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font = SUBHEADER_FONT
                cell.fill = SUBHEADER_FILL
                cell.border = THIN_BORDER
            row += 1

            for t in period_trades:
                vals = [
                    fmt_date(t.get("entry_date", "")),
                    fmt_date(t.get("exit_date", "")),
                    t.get("direction", ""),
                    t.get("shares", 0),
                    fmt_money(t.get("entry_price", 0)),
                    fmt_money(t.get("exit_price", 0)),
                    fmt_money(t.get("pnl", 0)),
                    f"{t.get('pnl_pct', 0)*100:.2f}%",
                    t.get("exit_reason", ""),
                ]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=c, value=v)
                    cell.font = DATA_FONT
                    cell.border = THIN_BORDER
                    if c in (5, 6, 7):
                        cell.number_format = "¥#,##0.00"
                row += 1
        else:
            ws.cell(row=row, column=1, value="（本期间无平仓交易）").font = Font(name="Arial", size=9, color="999999", italic=True)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 40
        ws.column_dimensions["J"].width = 12
        ws.column_dimensions["K"].width = 35

        # Freeze panes at the daily log header
        # (approximate — freeze first 2 columns)
        ws.freeze_panes = "C1"

    # ── Save ────────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"Excel saved: {output_path}")
    print(f"  Sheets: {len(wb.sheetnames)} (1 summary + {len(decisions)} period sheets)")


def _sheet_name(dec_date: str, decisions: list, idx: int) -> str:
    """Generate a sheet name like '01-05~01-25'."""
    start = dec_date[5:]  # MM-DD
    if idx + 1 < len(decisions):
        end = fmt_date(decisions[idx + 1]["date"])[5:]
    else:
        end = "结束"
    return f"{start}~{end}"


if __name__ == "__main__":
    result_file = sys.argv[1] if len(sys.argv) > 1 else \
        "backtest_results/hybrid_custom_000423/000423/result_20260707_093253.json"
    output_file = sys.argv[2] if len(sys.argv) > 2 else \
        "backtest_results/hybrid_custom_000423/000423/pm_rules_trading_log.xlsx"

    build_excel(result_file, output_file)
