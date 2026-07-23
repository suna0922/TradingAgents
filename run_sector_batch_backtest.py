#!/usr/bin/env python3
"""
Batch backtest: randomly pick 3 stocks per sector, run with industry-appropriate master presets.

Usage:
  .venv/bin/python run_sector_batch_backtest.py [--start 2026-01-01] [--end 2026-07-06] [--per-sector 3] [--workers 2]
"""

import sys, os, json, random, subprocess, time, shutil
from pathlib import Path
from datetime import datetime
from collections import defaultdict

_PROJECT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _PROJECT)

# ── Industry classification → preset mapping ──────────────────────
# baostock 证监会行业分类代码 → industry preset key
_SECTOR_MAP = {
    # 科技创新: 电子/计算机/通信/软件/互联网/电信
    "C39": "tech_innovation",  # 计算机、通信和其他电子设备制造业
    "I65": "tech_innovation",  # 软件和信息技术服务业
    "I64": "tech_innovation",  # 互联网和相关服务
    "I63": "tech_innovation",  # 电信、广播电视和卫星传输服务
    "C40": "tech_innovation",  # 仪器仪表制造业

    # 新能源: 电气机械(含光伏/锂电/储能)
    "C38": "new_energy",       # 电气机械和器材制造业

    # 消费白马: 酒/食品/零售/纺织/家具
    "C15": "consumer",         # 酒、饮料和精制茶制造业
    "C14": "consumer",         # 食品制造业
    "C13": "consumer",         # 农副食品加工业
    "F52": "consumer",         # 零售业
    "F51": "consumer",         # 批发业
    "C18": "consumer",         # 纺织服装、服饰业
    "C17": "consumer",         # 纺织业
    "C21": "consumer",         # 家具制造业
    "C24": "consumer",         # 文教、工美、体育和娱乐用品制造业

    # 医药医疗
    "C27": "pharma",           # 医药制造业
    "Q84": "pharma",           # 卫生
    "M73": "pharma",           # 研究和试验发展 (含CRO/CDMO)

    # 金融: 银行/券商/保险
    "J66": "finance",          # 货币金融服务 (银行)
    "J67": "finance",          # 资本市场服务 (券商)
    "J68": "finance",          # 保险业
    "J69": "finance",          # 其他金融业

    # 周期资源: 有色/黑色/化工/煤炭/石油/化纤
    "C32": "cyclical",         # 有色金属冶炼和压延加工业
    "C31": "cyclical",         # 黑色金属冶炼和压延加工业
    "C26": "cyclical",         # 化学原料和化学制品制造业
    "C28": "cyclical",         # 化学纤维制造业
    "B06": "cyclical",         # 煤炭开采和洗选业
    "B09": "cyclical",         # 有色金属矿采选业
    "B07": "cyclical",         # 石油和天然气开采业
    "C25": "cyclical",         # 石油、煤炭及其他燃料加工业
    "B08": "cyclical",         # 黑色金属矿采选业

    # 高端制造: 专用设备/通用设备/汽车/金属制品
    "C35": "manufacturing",    # 专用设备制造业
    "C34": "manufacturing",    # 通用设备制造业
    "C36": "manufacturing",    # 汽车制造业
    "C33": "manufacturing",    # 金属制品业
    "C41": "manufacturing",    # 其他制造业
    "C37": "manufacturing",    # 铁路、船舶、航空航天和其他运输设备制造业

    # 地产基建
    "K70": "real_estate",      # 房地产业
    "E48": "real_estate",      # 土木工程建筑业
    "E49": "real_estate",      # 建筑安装业
    "E50": "real_estate",      # 建筑装饰、装修和其他建筑业
    "E47": "real_estate",      # 房屋建筑业

    # 公用环保: 电力/燃气/水/环保/交运
    "D44": "utility",          # 电力、热力生产和供应业
    "D45": "utility",          # 燃气生产和供应业
    "D46": "utility",          # 水的生产和供应业
    "N77": "utility",          # 生态保护和环境治理业
    "N78": "utility",          # 公共设施管理业
    "N76": "utility",          # 水利管理业
    "G53": "utility",          # 铁路运输业
    "G54": "utility",          # 道路运输业
    "G55": "utility",          # 水上运输业
    "G56": "utility",          # 航空运输业
    "G58": "utility",          # 多式联运和运输代理业
    "G59": "utility",          # 装卸搬运和仓储业
    "G60": "utility",          # 邮政业

    # 农业养殖
    "A01": "agriculture",      # 农业
    "A02": "agriculture",      # 林业
    "A03": "agriculture",      # 畜牧业
    "A04": "agriculture",      # 渔业
    "A05": "agriculture",      # 农、林、牧、渔专业及辅助性活动
}

# 10 sector presets with display names
_SECTOR_INFO = {
    "tech_innovation": "科技创新",
    "new_energy":      "新能源",
    "consumer":        "消费白马",
    "pharma":          "医药医疗",
    "finance":         "金融",
    "cyclical":        "周期资源",
    "manufacturing":   "高端制造",
    "real_estate":     "地产基建",
    "utility":         "公用环保",
    "agriculture":     "农业养殖",
}


def fetch_sector_stocks():
    """Get all A-share stocks with sector classification from baostock."""
    import baostock as bs
    lg = bs.login()
    if lg.error_code != '0':
        raise RuntimeError(f"baostock login failed: {lg.error_msg}")

    rs = bs.query_stock_industry()
    if rs.error_code != '0':
        bs.logout()
        raise RuntimeError(f"query_stock_industry failed: {rs.error_msg}")

    all_data = []
    while (rs.error_code == '0') and rs.next():
        all_data.append(rs.get_row_data())
    bs.logout()

    # Group by sector
    sector_stocks = defaultdict(list)
    skipped_no_industry = 0
    skipped_st = 0
    skipped_unmapped = 0

    for row in all_data:
        code = row[1]          # e.g. "sh.600000"
        name = row[2]          # e.g. "浦发银行"
        industry = row[3]      # e.g. "J66货币金融服务"

        if not industry:
            skipped_no_industry += 1
            continue

        # Extract industry code (e.g. "J66" from "J66货币金融服务")
        ind_code = industry[:3] if len(industry) >= 3 else ""

        sector = _SECTOR_MAP.get(ind_code)
        if not sector:
            skipped_unmapped += 1
            continue

        # Filter: skip ST / *ST / 退 stocks
        if 'ST' in name or '退' in name:
            skipped_st += 1
            continue

        # Extract pure symbol (e.g. "600000" from "sh.600000")
        symbol = code.split('.')[1]

        # Skip non-mainboard stocks (keep 60xxxx, 00xxxx, 30xxxx)
        # 60xxxx = SH main board, 00xxxx = SZ main board, 30xxxx = ChiNext
        # Skip 68xxxx (STAR Market) for now - may have different trading rules
        # Skip 88xxxx (index), 50xxxx (fund), etc.
        if not (symbol.startswith('60') or symbol.startswith('00') or symbol.startswith('30')):
            continue

        sector_stocks[sector].append({"symbol": symbol, "name": name, "industry": industry})

    print(f"[SCREEN] Total stocks: {len(all_data)}")
    print(f"[SCREEN] Skipped (no industry): {skipped_no_industry}")
    print(f"[SCREEN] Skipped (ST/退): {skipped_st}")
    print(f"[SCREEN] Skipped (unmapped): {skipped_unmapped}")
    print(f"[SCREEN] Skipped (non-mainboard): {len(all_data) - skipped_no_industry - skipped_st - skipped_unmapped - sum(len(v) for v in sector_stocks.values())}")

    return dict(sector_stocks)


def pick_random_stocks(sector_stocks, per_sector=3, seed=42):
    """Randomly pick N stocks per sector."""
    random.seed(seed)
    picked = {}
    for sector, stocks in sorted(sector_stocks.items()):
        # Sort by symbol for determinism before random pick
        stocks_sorted = sorted(stocks, key=lambda x: x["symbol"])
        n = min(per_sector, len(stocks_sorted))
        chosen = random.sample(stocks_sorted, n)
        picked[sector] = chosen
        names = [f"{s['symbol']} {s['name']}" for s in chosen]
        print(f"  {_SECTOR_INFO.get(sector, sector):8s}: {', '.join(names)}")
    return picked


def run_single_backtest(symbol, name, sector, start_date, end_date, master_config):
    """Run a single backtest in a subprocess and return result path."""
    output_dir = f"backtest_results/batch_sector/{symbol}"
    os.makedirs(output_dir, exist_ok=True)

    # Build master config JSON for the script
    master_json = json.dumps(master_config, ensure_ascii=False)

    # Create a temporary per-stock runner script
    runner_script = f"""
import sys, os, json
sys.path.insert(0, {_PROJECT!r})

# Load .env
from dotenv import load_dotenv
load_dotenv()

# Apply industry master config
from tradingagents.default_config import DEFAULT_CONFIG
DEFAULT_CONFIG["master_config"] = json.loads('{master_json}')

# Backup & clear memory log
import pandas as pd
_mem = DEFAULT_CONFIG.get("memory_log_path", "")
if _mem and os.path.exists(_mem):
    import shutil
    _bak = _mem + ".bak_" + pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
    os.rename(_mem, _bak)

# Clear L1 cache for this symbol
_cache_dir = os.path.join({_PROJECT!r}, "backtest_results", "batch_sector", "l1_cache")
if os.path.exists(_cache_dir):
    for f in os.listdir(_cache_dir):
        if f.startswith("q_{symbol}_") and f.endswith(".json"):
            os.remove(os.path.join(_cache_dir, f))

# Run backtest
from backtest_hybrid import HybridBacktestEngine

engine = HybridBacktestEngine(
    symbol="{symbol}",
    start_date="{start_date}",
    end_date="{end_date}",
    initial_cash=1_000_000.0,
    price_change_threshold=0.10,
    stale_days=15,
    output_dir="{output_dir}",
)

result = engine.run()

# Save results
out_dir = os.path.join({_PROJECT!r}, "{output_dir}", "{symbol}")
os.makedirs(out_dir, exist_ok=True)
ts = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")

with open(os.path.join(out_dir, f"result_{{ts}}.json"), "w", encoding="utf-8") as f:
    json.dump(result.to_dict(), f, indent=2, ensure_ascii=False, default=str)
with open(os.path.join(out_dir, f"summary_{{ts}}.json"), "w", encoding="utf-8") as f:
    json.dump(result.summary, f, indent=2, ensure_ascii=False)

print(f"DONE: {symbol} return={{result.summary.get('total_return_pct', 0):.2f}}%")
"""

    script_path = os.path.join(output_dir, f"_runner_{symbol}.py")
    with open(script_path, "w") as f:
        f.write(runner_script)

    log_file = os.path.join(output_dir, f"backtest.log")

    print(f"\n[BACKTEST] {symbol} {name} | sector={_SECTOR_INFO.get(sector, sector)} | {start_date} ~ {end_date}")

    try:
        with open(log_file, "w") as logf:
            proc = subprocess.run(
                [".venv/bin/python", script_path],
                stdout=logf,
                stderr=subprocess.STDOUT,
                cwd=_PROJECT,
                timeout=3600,  # 1 hour max per stock
            )
        success = proc.returncode == 0
    except subprocess.TimeoutExpired:
        print(f"  [TIMEOUT] {symbol} exceeded 1 hour, skipping")
        success = False
    except Exception as e:
        print(f"  [ERROR] {symbol}: {e}")
        success = False

    # Read result
    result_data = None
    summary_data = None
    if success:
        out_dir = os.path.join(_PROJECT, output_dir, symbol)
        if os.path.exists(out_dir):
            for fname in sorted(os.listdir(out_dir), reverse=True):
                if fname.startswith("summary_") and fname.endswith(".json"):
                    with open(os.path.join(out_dir, fname)) as f:
                        summary_data = json.load(f)
                    break
            for fname in sorted(os.listdir(out_dir), reverse=True):
                if fname.startswith("result_") and fname.endswith(".json"):
                    with open(os.path.join(out_dir, fname)) as f:
                        result_data = json.load(f)
                    break

    return {
        "symbol": symbol,
        "name": name,
        "sector": sector,
        "sector_name": _SECTOR_INFO.get(sector, sector),
        "success": success,
        "summary": summary_data,
        "result": result_data,
    }


def generate_summary_excel(all_results, output_path):
    """Generate summary Excel with all backtest results."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: 总览 ──
    ws = wb.active
    ws.title = "回测总览"

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    sector_fills = {
        "tech_innovation": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "new_energy":      PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        "consumer":        PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "pharma":          PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
        "finance":         PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "cyclical":        PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
        "manufacturing":   PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid"),
        "real_estate":     PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid"),
        "utility":         PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        "agriculture":     PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
    }
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    red_font = Font(color="FF0000")
    green_font = Font(color="008000")

    # Title
    ws.merge_cells('A1:L1')
    ws['A1'] = f"板块批量回测总览 — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    headers = ["板块", "代码", "名称", "收益率", "基准收益", "超额收益", "最大回撤", "夏普比率", "交易次数", "胜率", "L1分析次数", "状态"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    row = 4
    for sector in sorted(all_results.keys()):
        for r in all_results[sector]:
            sector_fill = sector_fills.get(sector)
            s = r.get("summary") or {}

            values = [
                r.get("sector_name", ""),
                r.get("symbol", ""),
                r.get("name", ""),
                f"{s.get('total_return_pct', 0):+.2f}%" if r.get("success") else "N/A",
                f"{s.get('benchmark_return_pct', 0):+.2f}%" if r.get("success") else "N/A",
                f"{s.get('excess_return_pct', 0):+.2f}%" if r.get("success") else "N/A",
                f"{s.get('max_drawdown_pct', 0):.2f}%" if r.get("success") else "N/A",
                f"{s.get('sharpe_ratio', 0):.2f}" if r.get("success") else "N/A",
                s.get("total_trades", 0) if r.get("success") else 0,
                f"{s.get('win_rate', 0)*100:.0f}%" if r.get("success") and s.get('win_rate') else "N/A",
                s.get("total_l1_analyses", 0) if r.get("success") else 0,
                "成功" if r.get("success") else "失败",
            ]

            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                if sector_fill:
                    cell.fill = sector_fill
                # Color returns
                if col in (4, 6) and r.get("success"):
                    val = s.get('total_return_pct' if col == 4 else 'excess_return_pct', 0)
                    cell.font = red_font if val > 0 else green_font
                elif col == 7 and r.get("success"):
                    cell.font = red_font if s.get('max_drawdown_pct', 0) > 5 else None

            row += 1

    # Column widths
    widths = [10, 10, 12, 12, 12, 12, 12, 10, 10, 8, 12, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: 板块统计 ──
    ws2 = wb.create_sheet("板块统计")
    ws2.merge_cells('A1:G1')
    ws2['A1'] = "各板块平均表现统计"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].alignment = Alignment(horizontal='center')

    stat_headers = ["板块", "股票数", "平均收益率", "平均超额收益", "平均最大回撤", "平均夏普", "盈利比例"]
    for col, h in enumerate(stat_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row = 4
    for sector in sorted(all_results.keys()):
        results = [r for r in all_results[sector] if r.get("success") and r.get("summary")]
        if not results:
            continue

        n = len(results)
        avg_ret = sum(r["summary"].get("total_return_pct", 0) for r in results) / n
        avg_excess = sum(r["summary"].get("excess_return_pct", 0) for r in results) / n
        avg_dd = sum(r["summary"].get("max_drawdown_pct", 0) for r in results) / n
        avg_sharpe = sum(r["summary"].get("sharpe_ratio", 0) for r in results) / n
        win_count = sum(1 for r in results if r["summary"].get("total_return_pct", 0) > 0)
        win_ratio = win_count / n

        values = [
            _SECTOR_INFO.get(sector, sector),
            n,
            f"{avg_ret:+.2f}%",
            f"{avg_excess:+.2f}%",
            f"{avg_dd:.2f}%",
            f"{avg_sharpe:.2f}",
            f"{win_ratio*100:.0f}%",
        ]

        sector_fill = sector_fills.get(sector)
        for col, v in enumerate(values, 1):
            cell = ws2.cell(row=row, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if sector_fill:
                cell.fill = sector_fill
            if col == 3:
                cell.font = red_font if avg_ret > 0 else green_font
            elif col == 4:
                cell.font = red_font if avg_excess > 0 else green_font

        row += 1

    for i, w in enumerate([12, 10, 14, 14, 14, 10, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: 选股清单 ──
    ws3 = wb.create_sheet("选股清单")
    ws3.merge_cells('A1:E1')
    ws3['A1'] = "各板块随机选股清单"
    ws3['A1'].font = Font(bold=True, size=14)
    ws3['A1'].alignment = Alignment(horizontal='center')

    pick_headers = ["板块", "代码", "名称", "证监会行业", "大师配置"]
    for col, h in enumerate(pick_headers, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row = 4
    from tradingagents.masters.industry_presets import industry_preset
    for sector in sorted(all_results.keys()):
        preset = industry_preset(sector)
        preset_str = "; ".join(f"{k}={v}" for k, v in sorted(preset.items()))
        sector_fill = sector_fills.get(sector)
        for r in all_results[sector]:
            values = [
                _SECTOR_INFO.get(sector, sector),
                r.get("symbol", ""),
                r.get("name", ""),
                r.get("summary", {}).get("industry", "") if r.get("summary") else "",
                preset_str,
            ]
            for col, v in enumerate(values, 1):
                cell = ws3.cell(row=row, column=col, value=v)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center' if col < 5 else 'left')
                if sector_fill:
                    cell.fill = sector_fill
            row += 1

    for i, w in enumerate([12, 10, 12, 20, 60], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    print(f"\n[EXCEL] Summary saved to {output_path}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Batch sector backtest")
    parser.add_argument("--start", default="2026-01-01", help="Start date")
    parser.add_argument("--end", default="2026-07-06", help="End date")
    parser.add_argument("--per-sector", type=int, default=3, help="Stocks per sector")
    parser.add_argument("--seed", type=int, default=42, help="Random seed")
    args = parser.parse_args()

    # Adjust start date for weekends
    import pandas as pd
    start_dt = pd.Timestamp(args.start)
    if start_dt.dayofweek >= 5:
        start_dt = start_dt + pd.offsets.Day(7 - start_dt.dayofweek)
    start_date = start_dt.strftime("%Y-%m-%d")

    print(f"\n{'='*70}")
    print(f"  板块批量回测")
    print(f"  期间: {start_date} → {args.end}")
    print(f"  每板块: {args.per_sector} 支股票")
    print(f"  随机种子: {args.seed}")
    print(f"{'='*70}\n")

    # Step 1: Fetch stocks
    print("[STEP 1] Fetching A-share stocks from baostock...")
    sector_stocks = fetch_sector_stocks()

    print(f"\n[STEP 2] Sector distribution:")
    from tradingagents.masters.industry_presets import industry_preset, list_industries
    for sector in sorted(sector_stocks.keys()):
        print(f"  {_SECTOR_INFO.get(sector, sector):8s}: {len(sector_stocks[sector])} stocks")

    # Step 2: Pick random stocks
    print(f"\n[STEP 3] Randomly picking {args.per_sector} per sector (seed={args.seed}):")
    picked = pick_random_stocks(sector_stocks, per_sector=args.per_sector, seed=args.seed)

    total = sum(len(v) for v in picked.values())
    print(f"\n  Total stocks to backtest: {total}")

    # Step 3: Run backtests
    print(f"\n[STEP 4] Running backtests ({total} stocks)...")
    from tradingagents.masters.industry_presets import industry_preset

    all_results = {}
    completed = 0
    start_time = time.time()

    for sector in sorted(picked.keys()):
        all_results[sector] = []
        master_config = industry_preset(sector)

        for stock in picked[sector]:
            symbol = stock["symbol"]
            name = stock["name"]

            print(f"\n{'─'*60}")
            print(f"  [{completed+1}/{total}] {symbol} {name} | {_SECTOR_INFO.get(sector, sector)}")
            print(f"  Elapsed: {(time.time()-start_time)/60:.1f} min")
            print(f"{'─'*60}")

            result = run_single_backtest(
                symbol, name, sector,
                start_date, args.end,
                master_config
            )
            all_results[sector].append(result)
            completed += 1

            s = result.get("summary") or {}
            if result["success"]:
                print(f"  ✅ {symbol}: return={s.get('total_return_pct', 0):+.2f}% | excess={s.get('excess_return_pct', 0):+.2f}% | sharpe={s.get('sharpe_ratio', 0):.2f}")
            else:
                print(f"  ❌ {symbol}: FAILED")

    elapsed = (time.time() - start_time) / 60
    print(f"\n{'='*70}")
    print(f"  All backtests completed in {elapsed:.1f} minutes")
    print(f"  Success: {sum(1 for s in all_results.values() for r in s if r['success'])}/{total}")
    print(f"{'='*70}\n")

    # Step 4: Generate summary
    print("[STEP 5] Generating summary Excel...")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(_PROJECT, "backtest_results", "batch_sector", f"sector_batch_summary_{ts}.xlsx")
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    generate_summary_excel(all_results, excel_path)

    # Also save raw results JSON
    raw_path = os.path.join(_PROJECT, "backtest_results", "batch_sector", f"sector_batch_raw_{ts}.json")
    raw_data = {}
    for sector, results in all_results.items():
        raw_data[sector] = []
        for r in results:
            raw_data[sector].append({
                "symbol": r["symbol"],
                "name": r["name"],
                "sector": r["sector"],
                "success": r["success"],
                "summary": r["summary"],
            })
    with open(raw_path, "w", encoding="utf-8") as f:
        json.dump(raw_data, f, indent=2, ensure_ascii=False, default=str)
    print(f"[JSON] Raw results saved to {raw_path}")

    # Print final summary table
    print(f"\n{'='*70}")
    print(f"  Final Summary")
    print(f"{'='*70}")
    print(f"  {'板块':8s} | {'代码':8s} | {'名称':8s} | {'收益率':>8s} | {'超额':>8s} | {'夏普':>6s} | {'状态'}")
    print(f"  {'─'*70}")
    for sector in sorted(all_results.keys()):
        for r in all_results[sector]:
            s = r.get("summary") or {}
            ret = f"{s.get('total_return_pct', 0):+.2f}%" if r["success"] else "N/A"
            exc = f"{s.get('excess_return_pct', 0):+.2f}%" if r["success"] else "N/A"
            sha = f"{s.get('sharpe_ratio', 0):.2f}" if r["success"] else "N/A"
            status = "✅" if r["success"] else "❌"
            print(f"  {_SECTOR_INFO.get(sector,''):8s} | {r['symbol']:8s} | {r['name']:8s} | {ret:>8s} | {exc:>8s} | {sha:>6s} | {status}")
    print(f"  {'─'*70}")


if __name__ == "__main__":
    main()
