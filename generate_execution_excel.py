"""从回测缓存产物生成逐日执行详情 Excel（3-Sheet，按决策区间分段）。

不调用 LLM，纯本地计算。读取：
- backtest_results/cache/decisions/*.json → 3 个决策的 trading_rules
- backtest_results/cache/ohlcv/000423.csv → OHLCV 原始数据
- 实时计算技术指标 + 200 日均线

逐日评估每条规则的触发状态，输出同款 Excel。
"""

import sys, os, json, logging
from typing import Dict, List, Any, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from backtest.data_layer import DataLayer

logging.basicConfig(level=logging.WARNING, format="%(levelname)s %(name)s: %(message)s")


# ════════════════════════════════════════════════════════════════
# 工具函数
# ════════════════════════════════════════════════════════════════

def load_cache_decisions(cache_dir: str) -> List[Dict]:
    """加载所有缓存决策，按日期排序。"""
    import glob
    files = sorted(glob.glob(os.path.join(cache_dir, "*.json")))
    decisions = []
    for fp in files:
        with open(fp) as f:
            d = json.load(f)
        d["_file_date"] = os.path.basename(fp).replace("000423_", "").replace(".json", "")
        decisions.append(d)
    decisions.sort(key=lambda x: x["_file_date"])
    return decisions


def load_and_prepare_data(symbol: str, start: str, end: str):
    """加载 OHLCV + 计算技术指标 + 200日均线。"""
    data_start = "2024-01-01"
    dl = DataLayer(symbol=symbol, start_date=data_start, end_date=end)
    df = dl.fetch_ohlcv()
    df = dl.compute_indicators()

    pdf = pd.DataFrame(df.reset_index())
    pdf.columns = ["date"] + [c for c in df.columns]
    pdf["date_str"] = pdf["date"].astype(str).str[:10]
    pdf["close_200_sma"] = pdf["close"].rolling(200, min_periods=1).mean()

    mask = (pdf["date_str"] >= start) & (pdf["date_str"] <= end)
    pdf_bt = pdf[mask].copy().reset_index(drop=True)
    return pdf, pdf_bt


def build_decision_schedule(decisions: List[Dict]) -> List[Dict]:
    """构建决策时间表：每个决策的生效日期区间 + 规则集。

    Returns:
        [{start_date, end_date, decision, rules, label}, ...]
    """
    schedule = []
    for i, dec in enumerate(decisions):
        start = dec["_file_date"][:10]
        if i + 1 < len(decisions):
            end = decisions[i + 1]["_file_date"][:10]
        else:
            end = "2099-12-31"
        rules = dec.get("trading_rules", [])
        schedule.append({
            "start_date": start,
            "end_date": end,
            "decision": dec,
            "rules": rules,
            "label": f"决策{i+1} ({start} ~ {end if end != '2099-12-31' else '结束'})",
        })
    return schedule


def evaluate_rule_conditions(rule: Dict, row_dict: Dict) -> Dict:
    """评估单条规则的所有条件。"""
    conds = rule.get("conditions", [])
    cond_strs = []
    all_ok = True

    for c in conds:
        field = c.get("field", "")
        op = c.get("op", "")
        value = c.get("value", 0)
        actual = row_dict.get(field)

        if actual is None:
            all_ok = False
            cond_strs.append(f"{field}=N/A {op} {value} ❌")
            continue

        try:
            actual_f = float(actual)
            value_f = float(value)

            if op == "<=":
                ok = actual_f <= value_f
            elif op == "<":
                ok = actual_f < value_f
            elif op == ">=":
                ok = actual_f >= value_f
            elif op == ">":
                ok = actual_f > value_f
            elif op == "==":
                ok = abs(actual_f - value_f) < 0.0001
            elif op == "!=":
                ok = abs(actual_f - value_f) >= 0.0001
            else:
                ok = False

            emoji = "✅" if ok else "❌"
            cond_strs.append(f"{field}={actual_f:.4f} {op} {value} {emoji}")
            if not ok:
                all_ok = False
        except (ValueError, TypeError):
            all_ok = False
            cond_strs.append(f"{field}={actual} ERR")

    return {
        "triggered": all_ok,
        "conds_detail": " | ".join(cond_strs) if cond_strs else "无条件",
        "action": rule.get("action", "hold"),
    }


def get_effective_decision(date_str: str, schedule: List[Dict]) -> Dict:
    """根据日期找到生效的决策区间。"""
    for s in schedule:
        if s["start_date"] <= date_str < s["end_date"]:
            return s
    return None


# ════════════════════════════════════════════════════════════════
# 主流程
# ════════════════════════════════════════════════════════════════

def generate(
    symbol: str = "000423",
    start: str = "2025-01-01",
    end: str = "2025-03-01",
    output_path: str = "backtest_results/execution_daily_000423_2025-01-01_2025-03-01.xlsx",
):
    print("=" * 60)
    print(f"📊 生成回测执行详情 Excel（按决策区间分段）")
    print(f"   标的: {symbol} | 期间: {start} ~ {end}")
    print("=" * 60)

    # ── Step 1: 加载缓存决策 ──────────────────────────────────
    cache_dir = "backtest_results/cache/decisions"
    decisions = load_cache_decisions(cache_dir)
    print(f"\n📋 加载 {len(decisions)} 个决策:")
    for dec in decisions:
        rules = dec.get("trading_rules", [])
        print(f"   {dec['_file_date'][:10]} | {dec['direction']} | "
              f"pos={dec.get('position_pct', 0)} | {len(rules)} rules")

    schedule = build_decision_schedule(decisions)
    print(f"\n📋 决策区间:")
    for s in schedule:
        print(f"   {s['label']}: {len(s['rules'])} rules, "
              f"dir={s['decision']['direction']}, pos={s['decision'].get('position_pct', 0)}")

    # ── Step 2: 加载数据 ──────────────────────────────────────
    print(f"\n📋 加载数据...")
    full_pdf, bt_pdf = load_and_prepare_data(symbol, start, end)
    print(f"   完整数据: {len(full_pdf)} 行 | 回测期间: {len(bt_pdf)} 行")
    print(f"   日期范围: {bt_pdf['date_str'].iloc[0]} ~ {bt_pdf['date_str'].iloc[-1]}")

    # ── Step 3: 逐日执行规则匹配（按决策区间分段）─────────────
    print(f"\n📋 逐日评估规则触发 ({len(bt_pdf)} 个交易日)...")

    # 为每个区间收集独立的行数据
    # section_rows[section_idx] = list of row dicts
    section_rows: Dict[int, List[Dict]] = {}
    for i in range(len(schedule)):
        section_rows[i] = []

    for idx in range(len(bt_pdf)):
        row = bt_pdf.iloc[idx]
        date_str = row["date_str"]
        close = float(row["close"])

        eff = get_effective_decision(date_str, schedule)
        if eff is None:
            continue

        # 找到属于哪个区间
        section_idx = schedule.index(eff)
        rules = eff["rules"]

        # 构建 row_dict
        ma50 = float(row.get("close_50_sma", 0))
        ma200_val = float(row.get("close_200_sma", 0))
        macd_val = float(row.get("macd", 0))
        macds_val = float(row.get("macds", 0))
        macdh_val = float(row.get("macdh", 0))

        row_dict = {
            "close": close,
            "open": float(row["open"]),
            "high": float(row["high"]),
            "low": float(row["low"]),
            "volume": float(row["volume"]),
            "pct_chg": float(row.get("pct_chg", 0)),
            "close_50_sma": ma50,
            "close_200_sma": ma200_val,
            "macd": macd_val,
            "macds": macds_val,
            "macdh": macdh_val,
            "rsi": float(row.get("rsi", 0)),
        }

        # 逐条规则评估
        rule_details = {}
        for i, rule in enumerate(rules):
            rid = f"R{i+1}"
            result = evaluate_rule_conditions(rule, row_dict)
            rule_details[rid] = result

        triggered_list = [rid for rid, d in rule_details.items() if d["triggered"]]

        # 技术指标格式化
        cross_below_ma50 = "⚠️跌破" if (ma50 > 0 and close < ma50) else "✅"
        cross_below_ma200 = "⚠️跌破" if (ma200_val > 0 and close < ma200_val) else "✅"
        macd_death_cross = "⚠️是" if (macd_val > 0 and macd_val < macds_val) else "—"

        # 动作
        triggered_rules = [r for i, r in enumerate(rules) if rule_details[f"R{i+1}"]["triggered"]]
        if triggered_rules:
            triggered_rules.sort(key=lambda r: r.get("priority", 0), reverse=True)
            action = triggered_rules[0].get("action", "hold")
        else:
            action = "HOLD"

        base = {
            "日期": date_str,
            "收盘价": round(close, 2),
            "涨跌幅%": round(float(row.get("pct_chg", 0)), 2),
            "50日均线": round(ma50, 2),
            "200日均线": round(ma200_val, 2),
            "收盘<50日": cross_below_ma50,
            "收盘<200日": cross_below_ma200,
            "MACD": round(macd_val, 4),
            "MACD柱": round(macdh_val, 4),
            "MACD信号": round(macds_val, 4),
            "MACD零轴死叉": macd_death_cross,
        }

        # 规则触发列（按该区间的规则）
        rule_cols = {}
        cond_cols = {}
        for i in range(len(rules)):
            rid = f"R{i+1}"
            rule_cols[f"{rid}_触发"] = "🔥是" if rule_details[rid]["triggered"] else "—"
            cond_cols[f"{rid}_条件"] = rule_details[rid]["conds_detail"]

        extra = {
            "触发规则": ", ".join(triggered_list) if triggered_list else "—",
            "触发数": len(triggered_list),
            "动作": action,
            "区间": eff["label"],
        }

        row_data = {**base, **rule_cols, **extra, **cond_cols, "_section_idx": section_idx}
        section_rows[section_idx].append(row_data)

    total_rows = sum(len(v) for v in section_rows.values())
    print(f"   ✅ 收集到 {total_rows} 天的数据")

    for si, rows in section_rows.items():
        triggered_days = [r for r in rows if r["触发数"] > 0]
        print(f"   📍 {schedule[si]['label']}: {len(rows)} 天, 触发 {len(triggered_days)} 天")
        for td in triggered_days[:5]:
            print(f"       {td['日期']}: {td['触发规则']} → 动作={td['动作']}")
        if len(triggered_days) > 5:
            print(f"       ... 还有 {len(triggered_days) - 5} 天")

    # ── Step 4: 生成 Excel ────────────────────────────────────
    print(f"\n📋 生成 Excel: {output_path}")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()

    # 颜色定义
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    triggered_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    triggered_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    action_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    normal_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    sep_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    base_cols = ["日期", "收盘价", "涨跌幅%", "50日均线", "200日均线",
                 "收盘<50日", "收盘<200日", "MACD", "MACD柱", "MACD信号",
                 "MACD零轴死叉"]

    # ── Sheet 1: 每日执行（按区间分段）─────────────────────────
    ws1 = wb.active
    ws1.title = "每日执行"

    current_row = 1

    for si in range(len(schedule)):
        s = schedule[si]
        rows = section_rows[si]
        rules = s["rules"]
        n_rules = len(rules)

        if n_rules == 0:
            continue

        # 区间标题行
        cell = ws1.cell(row=current_row, column=1, value=s["label"])
        cell.fill = section_fill
        cell.font = section_font
        cell.border = thin_border
        # 合并标题行
        total_cols = len(base_cols) + n_rules + 4 + n_rules  # base + R触发 + extra(触发规则,触发数,动作,区间) + R条件
        for c in range(2, total_cols + 1):
            cell = ws1.cell(row=current_row, column=c, value="")
            cell.fill = section_fill
            cell.border = thin_border
        ws1.merge_cells(start_row=current_row, start_column=1,
                        end_row=current_row, end_column=total_cols)
        current_row += 1

        # 列头
        rule_trigger_cols = [f"R{i+1}_触发" for i in range(n_rules)]
        extra_cols = ["触发规则", "触发数", "动作", "区间"]
        rule_cond_cols = [f"R{i+1}_条件" for i in range(n_rules)]
        all_cols = base_cols + rule_trigger_cols + extra_cols + rule_cond_cols

        for col_idx, col_name in enumerate(all_cols, 1):
            cell = ws1.cell(row=current_row, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        current_row += 1

        # 数据行
        for row_data in rows:
            for col_idx, col_name in enumerate(all_cols, 1):
                val = row_data.get(col_name, "")
                cell = ws1.cell(row=current_row, column=col_idx, value=val)
                cell.font = normal_font
                cell.border = thin_border

                if col_name.endswith("_触发") and val == "🔥是":
                    cell.fill = triggered_fill
                    cell.font = triggered_font
                if col_name == "动作" and val != "HOLD" and val != "hold":
                    cell.fill = action_fill
                    cell.font = Font(name="微软雅黑", size=10, bold=True)
            current_row += 1

        # 区间间空一行
        current_row += 1

    ws1.freeze_panes = "A2"
    ws1.column_dimensions["A"].width = 12
    for c in range(2, 50):
        ws1.column_dimensions[get_column_letter(c)].width = 14
    # 条件详情列加宽
    # (条件详情列在最右侧，需要根据实际列数调整)

    # ── Sheet 2: 规则分析（按区间分段）─────────────────────────
    ws2 = wb.create_sheet("规则分析")

    analysis_cols = ["规则", "动作", "优先级", "条件数", "条件详情", "触发天数", "首次触发"]
    current_row = 1

    for si in range(len(schedule)):
        s = schedule[si]
        rules = s["rules"]
        rows = section_rows[si]

        if not rules:
            continue

        # 区间标题
        cell = ws2.cell(row=current_row, column=1, value=s["label"])
        cell.fill = section_fill
        cell.font = section_font
        cell.border = thin_border
        for c in range(2, len(analysis_cols) + 1):
            cell = ws2.cell(row=current_row, column=c, value="")
            cell.fill = section_fill
            cell.border = thin_border
        ws2.merge_cells(start_row=current_row, start_column=1,
                        end_row=current_row, end_column=len(analysis_cols))
        current_row += 1

        # 列头
        for col_idx, col_name in enumerate(analysis_cols, 1):
            cell = ws2.cell(row=current_row, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        current_row += 1

        # 规则行
        for i, rule in enumerate(rules):
            rid = f"R{i+1}"
            conds = rule.get("conditions", [])
            cond_strs = []
            for c in conds:
                cond_strs.append(f"{c.get('field')} {c.get('op')} {c.get('value')} [{c.get('source_text', '')[:30]}]")

            # 统计该规则在区间内的触发情况
            trigger_col = f"{rid}_触发"
            triggered_days = [r for r in rows if r.get(trigger_col) == "🔥是"]
            triggered_count = len(triggered_days)
            first_trigger = triggered_days[0]["日期"] if triggered_days else "—"

            row_data = [
                f"{rid} {rule.get('name', '?')[:40]}",
                rule.get("action", "?"),
                rule.get("priority", 0),
                len(conds),
                " AND ".join(cond_strs),
                triggered_count,
                first_trigger,
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws2.cell(row=current_row, column=col_idx, value=val)
                cell.font = normal_font
                cell.border = thin_border
            current_row += 1

        current_row += 1  # 区间间空一行

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["E"].width = 80
    ws2.freeze_panes = "A2"

    # ── Sheet 3: 条件详情（按区间分段）─────────────────────────
    ws3 = wb.create_sheet("条件详情")
    cond_detail_cols = ["规则", "条件序号", "字段", "操作符", "阈值", "原文描述"]
    current_row = 1

    for si in range(len(schedule)):
        s = schedule[si]
        rules = s["rules"]

        if not rules:
            continue

        # 区间标题
        cell = ws3.cell(row=current_row, column=1, value=s["label"])
        cell.fill = section_fill
        cell.font = section_font
        cell.border = thin_border
        for c in range(2, len(cond_detail_cols) + 1):
            cell = ws3.cell(row=current_row, column=c, value="")
            cell.fill = section_fill
            cell.border = thin_border
        ws3.merge_cells(start_row=current_row, start_column=1,
                        end_row=current_row, end_column=len(cond_detail_cols))
        current_row += 1

        # 列头
        for col_idx, col_name in enumerate(cond_detail_cols, 1):
            cell = ws3.cell(row=current_row, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        current_row += 1

        # 条件行
        for i, rule in enumerate(rules):
            rid = f"R{i+1}"
            conds = rule.get("conditions", [])
            for j, c in enumerate(conds):
                row_data = [
                    f"{rid} {rule.get('name', '?')[:25]}",
                    j + 1,
                    c.get("field", ""),
                    c.get("op", ""),
                    c.get("value", ""),
                    c.get("source_text", "")[:80],
                ]
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws3.cell(row=current_row, column=col_idx, value=val)
                    cell.font = normal_font
                    cell.border = thin_border
                current_row += 1

        current_row += 1  # 区间间空一行

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["F"].width = 60
    ws3.freeze_panes = "A2"

    wb.save(output_path)
    print(f"   ✅ Excel 已保存: {output_path}")
    print(f"   Sheets: 每日执行（{len(schedule)} 区间分段） | 规则分析（{len(schedule)} 区间分段） | 条件详情（{len(schedule)} 区间分段）")

    return output_path


if __name__ == "__main__":
    output = generate()
    print(f"\n🎉 完成！输出文件: {output}")
