"""测试执行引擎日频规则匹配 — 东阿阿胶 000423, 2025-02-01 ~ 2025-03-01

直接用现有 ExecutionEngine + BacktestConfig + 数据加载 + PM 规则解析，
逐日调用 execute()，收集每条规则的触发情况，输出 Excel。
"""

import sys, os, logging, json
from datetime import datetime
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from backtest.models import (
    BacktestConfig, PortfolioState, WeeklyDecision, TradeDirection,
    PriceCondition, TechnicalTriggers, FundamentalGuards,
)
from backtest.data_layer import DataLayer
from backtest.execution_engine import ExecutionEngine
from backtest.trading_rules import RuleParser, TradingRule

logging.basicConfig(level=logging.WARNING, format="%(levelname)s %(name)s: %(message)s")

# ── 配置 ────────────────────────────────────────────────────────
SYMBOL = "000423"
DATA_START = "2024-01-01"   # 需要足够历史数据计算200日均线
START_DATE = "2025-02-01"    # 实际回测开始
END_DATE = "2025-03-01"      # 回测结束
INITIAL_CASH = 1_000_000

# ── PM 原始输出 ─────────────────────────────────────────────────
PM_OUTPUT = """
Rating: Underweight

Executive Summary:
采纳研究经理Underweight建议，在56元附近启动减仓，目标将仓位降至标准权重的60%-70%（即减仓30%-40%）。分两阶段执行：第一阶段在56元附近减持目标量的一半；若收盘价有效跌破50日均线（55.65元），执行第二阶段剩余减持。硬止损红线为收盘价有效跌破200日均线（53.96元），触发即清仓。时间框架3-6个月，等待2026年中报验证OCF/净利润是否企稳。

Investment Thesis:
Bear方核心证据——OCF/归母净利润五年连降79%（年报2021：6.36倍→年报2025：1.32倍）、股息发放率124.05%（年报2024）与94.04%（年报2025）、权益四年零增长（年报2022：103.51亿元→年报2025：103.63亿元）——构成了盈利质量趋势性恶化的清晰图景，且连续两年Q1经营现金流近乎枯竭（季报2025Q1：-0.71倍；季报2026Q1：0.06倍）印证了这一趋势。但Bull方的三大安全垫——73.47%毛利率（年报2025）、0.16%有息负债率（年报2025）、52.67亿元在手现金（年报2025）——使公司远未触及生存危机，不适用Sell评级。激进派将"管理效率低下"等同于"价值毁灭"过度简化了现实；保守派对趋势恶化信号过于迟钝。Underweight在承认风险与管理风险之间取得最佳平衡：减仓30%-40%以降低下行敞口，保留核心仓位以捕捉品牌价值回归潜力。技术面上MACD动能衰减97.5%（1.254→0.031）是短期警示，但200日均线（53.96元）和50日均线（55.65元）的稳定上行提供了中期结构支撑。

Time Horizon: 3-6个月

Trading Rules:

 • [RULE1] [stop_loss] IF 收盘价有效跌破200日均线 THEN stop_loss — 无条件清仓，仓位降至零 (@ 53.96元 (200日均线))
 • [RULE2] [reduce_position] IF 收盘价有效跌破50日均线 THEN reduce_position — 执行第二阶段减持，减持剩余目标减仓量（总仓位11%-20%），总仓位降至标准权重的60%-70% (@ 55.65元 (50日均线))
 • [RULE3] [entry_zone] IF 股价触及56元附近 THEN reduce_position — 执行第一阶段减持，减持目标减仓量的一半（总仓位15%-20%） (@ 56.0元 (当前价格区间))
 • [RULE4] [observation_anchor] IF 季报OCF/归母净利润连续两个季度低于1.0倍 THEN alert_only — 重新评估核心仓位安全边际，考虑进一步下调至Sell (@ N/A)
 • [RULE5] [observation_anchor] IF 季报2026Q2 OCF/归母净利润回升至2.0倍以上 THEN alert_only — 评估是否回补部分仓位，上调评级至Hold (@ N/A)
 • [RULE6] [observation_anchor] IF 年报2026股息发放率再次超过100% THEN alert_only — 视为管理层战略迷茫的强烈信号，加速减持节奏 (@ N/A)
 • [RULE7] [observation_anchor] IF MACD在零轴上方形成死叉 THEN downgrade — 加速减持节奏，提前执行第二阶段减仓 (@ N/A (MACD零轴))
"""

# ── FA 指标（PM 生成时的基本面数据，2024年报 + 2025Q1） ────────
FA_METRICS = {
    "annual_ocf_to_netprofit": 1.32,      # PM: 年报2025: 1.32倍
    "annual_dividend_payout": 94.04,       # PM: 年报2025: 94.04%
    "annual_gross_margin": 73.47,          # PM: 73.47%
    "annual_interest_debt": 0.16,          # PM: 0.16%有息负债率
    "annual_cash_billion": 52.67,          # PM: 52.67亿元
    "quarter_ocf_to_netprofit": -0.71,     # PM: 季报2025Q1: -0.71倍
    "quarter_gross_margin": 72.0,          # 估算
    "quarter_net_margin": 18.0,            # 估算
}


def main():
    print("=" * 60)
    print(f"🧪 日执行引擎规则匹配测试")
    print(f"   标的: {SYMBOL} | 期间: {START_DATE} ~ {END_DATE}")
    print("=" * 60)

    # ── 1. 手动构建7条规则 ────────────────────────────────────────
    # 注意：Condition.evaluate() 不支持跨字段比较（如 close <= close_50_sma），
    # 因此跨字段条件通过 row_dict 中预先计算的派生字段实现。
    print("\n📋 Step 1: 手动构建 PM 规则")
    from backtest.trading_rules import ComparisonOp, Condition, RuleAction
    
    rules = [
        # RULE1: 收盘价有效跌破200日均线 → 止损清仓
        # 使用派生字段 _cross_below_ma200（close < close_200_sma 时为 1）
        TradingRule(
            name="RULE1 stop_loss 跌破200日线",
            action=RuleAction.STOP_LOSS,
            conditions=[
                Condition(field="_cross_below_ma200", op=ComparisonOp.GT, value=0,
                          source_text="收盘价有效跌破200日均线"),
            ],
            priority=95,
            source_sentence="IF 收盘价有效跌破200日均线 THEN stop_loss (@ 53.96元)",
        ),
        # RULE2: 收盘价有效跌破50日均线 → 减仓
        TradingRule(
            name="RULE2 reduce_position 跌破50日线",
            action=RuleAction.SELL_HALF,
            conditions=[
                Condition(field="_cross_below_ma50", op=ComparisonOp.GT, value=0,
                          source_text="收盘价有效跌破50日均线"),
            ],
            priority=85,
            source_sentence="IF 收盘价有效跌破50日均线 THEN reduce_position (@ 55.65元)",
        ),
        # RULE3: 股价触及56元附近（55~57）→ 第一阶段减持
        TradingRule(
            name="RULE3 reduce_position 触及56元",
            action=RuleAction.SELL_PCT,
            conditions=[
                Condition(field="close", op=ComparisonOp.GTE, value=55.0,
                          source_text="股价触及56元附近(下界)"),
                Condition(field="close", op=ComparisonOp.LTE, value=57.0,
                          source_text="股价触及56元附近(上界)"),
            ],
            priority=70,
            pct=0.5,
            source_sentence="IF 股价触及56元附近 THEN reduce_position 减持一半",
        ),
        # RULE4: 季报OCF/净利润 < 1.0倍 → alert
        TradingRule(
            name="RULE4 alert 季报OCF连续低",
            action=RuleAction.HOLD,
            conditions=[
                Condition(field="quarter_ocf_to_netprofit", op=ComparisonOp.LT, value=1.0,
                          source_text="季报OCF/净利润 < 1.0倍"),
            ],
            priority=55,
            source_sentence="IF 季报OCF/净利润连续两季度低于1.0倍 THEN alert_only",
        ),
        # RULE5: 季报OCF/净利润 > 2.0倍 → alert
        TradingRule(
            name="RULE5 alert 季报OCF回升",
            action=RuleAction.HOLD,
            conditions=[
                Condition(field="quarter_ocf_to_netprofit", op=ComparisonOp.GT, value=2.0,
                          source_text="季报OCF/净利润 > 2.0倍"),
            ],
            priority=50,
            source_sentence="IF 季报2026Q2 OCF/净利润回升至2.0倍以上 THEN alert_only",
        ),
        # RULE6: 年报股息发放率 > 100% → alert
        TradingRule(
            name="RULE6 alert 股息发放率超100%",
            action=RuleAction.HOLD,
            conditions=[
                Condition(field="annual_dividend_payout", op=ComparisonOp.GT, value=100.0,
                          source_text="年报股息发放率 > 100%"),
            ],
            priority=50,
            source_sentence="IF 年报2026股息发放率再次超过100% THEN alert_only",
        ),
        # RULE7: MACD在零轴上方形成死叉 → downgrade
        # 使用派生字段 _macd_death_cross（MACD>0 且 MACD<=MACD信号线 时为 1）
        TradingRule(
            name="RULE7 downgrade MACD死叉",
            action=RuleAction.SELL_PCT,
            conditions=[
                Condition(field="_macd_death_cross", op=ComparisonOp.GT, value=0,
                          source_text="MACD在零轴上方形成死叉"),
            ],
            priority=60,
            pct=0.5,
            source_sentence="IF MACD在零轴上方形成死叉 THEN downgrade",
        ),
    ]
    print(f"   手动构建 {len(rules)} 条规则:")
    for r in sorted(rules, key=lambda x: x.priority, reverse=True):
        print(f"   [{r.name}] pri={r.priority} action={r.action.value} "
              f"conds={len(r.conditions)} enabled={r.enabled}")
        for c in r.conditions:
            print(f"      {c.field} {c.op.value} {c.value}")

    # ── 2. 构建 WeeklyDecision ───────────────────────────────────
    print("\n📋 Step 2: 构建 WeeklyDecision")
    decision = WeeklyDecision(
        direction=TradeDirection.SELL,  # Underweight → SELL
        position_pct=0.30,
        price_cond=PriceCondition(stop_loss=53.96),
        technical_triggers=TechnicalTriggers(),
        fundamental_guards=FundamentalGuards(),
        decision_date=START_DATE,
        signal_raw="Underweight",
        pm_rating="Underweight",
        pm_raw_output=PM_OUTPUT,
        trading_rules=rules,
        rules_parsed_ok=len(rules) > 0,
    )
    print(f"   direction={decision.direction.value} position_pct={decision.position_pct}")
    print(f"   rules_parsed_ok={decision.rules_parsed_ok} rules={len(decision.trading_rules)}")

    # ── 3. 加载数据 ──────────────────────────────────────────────
    print("\n📋 Step 3: 加载数据")
    config = BacktestConfig(
        symbol=SYMBOL, start_date=START_DATE, end_date=END_DATE,
        initial_cash=INITIAL_CASH,
    )
    data_layer = DataLayer(symbol=SYMBOL, start_date=DATA_START, end_date=END_DATE)
    df_full = data_layer.fetch_ohlcv()
    print(f"   原始数据: {len(df_full)} 行 ({DATA_START}~{END_DATE})")
    df_full = data_layer.compute_indicators()
    df_full = pd.DataFrame(df_full.reset_index())
    df_full["date"] = df_full["date"].astype(str).str[:10]
    # 在完整数据上计算 200 日均线
    df_full["close_200_sma"] = df_full["close"].rolling(200, min_periods=1).mean()
    # 截取实际回测期间
    df = df_full[df_full["date"] >= START_DATE].copy()
    df = df.reset_index(drop=True)
    print(f"   回测期间: {len(df)} 行 ({START_DATE}~{END_DATE}), 列: {list(df.columns)}")

    # ── 4. 初始化执行引擎 ─────────────────────────────────────────
    print("\n📋 Step 4: 初始化执行引擎")
    engine = ExecutionEngine(config, data_layer)
    portfolio = PortfolioState(cash=INITIAL_CASH)

    # 先满仓买入（模拟已有持仓，这样减仓规则才有意义）
    first_close = float(df.iloc[0]["close"])
    first_shares = int(INITIAL_CASH / first_close / 100) * 100
    cost = first_shares * first_close
    commission = max(cost * config.commission_rate, config.min_commission)
    portfolio.cash -= (cost + commission)
    portfolio.shares = first_shares
    print(f"   初始建仓: {first_shares} 股 @ {first_close:.2f} "
          f"(成本 ¥{cost+commission:,.0f}, 现金 ¥{portfolio.cash:,.0f})")

    # ── 5. 逐日执行并收集规则触发详情 ─────────────────────────────
    print(f"\n📋 Step 5: 逐日执行 ({len(df)} 个交易日)")
    
    rows = []
    for idx in range(len(df)):
        row = df.iloc[idx]
        date_str = str(row["date"])[:10]
        close = float(row["close"])

        # 构建 row_dict（模拟执行引擎中的逻辑）
        row_dict = row.to_dict()
        row_dict["_close"] = close
        row_dict["_high"] = float(row.get("high", 0))
        row_dict["_low"] = float(row.get("low", 0))
        row_dict["close"] = close
        row_dict.update(FA_METRICS)

        # 计算跨字段派生条件
        ma50 = float(row_dict.get("close_50_sma", 0))
        ma200 = float(row_dict.get("close_200_sma", 0)) if "close_200_sma" in row_dict else 0
        if ma200 == 0:
            # 从完整历史数据中计算（df 是回测期间的，但我们需要更长历史）
            # 简单方案：从完整原始数据中取
            pass
        row_dict["close_200_sma"] = ma200
        macd_val = float(row_dict.get("macd", 0))
        macds_val = float(row_dict.get("macds", 0))
        
        # 将派生字段注入 fa_metrics（执行引擎会把 fa_metrics 注入 row_dict）
        daily_fa = dict(FA_METRICS)
        daily_fa["close_200_sma"] = ma200
        daily_fa["_cross_below_ma50"] = 1 if (ma50 > 0 and close < ma50) else 0
        daily_fa["_cross_below_ma200"] = 1 if (ma200 > 0 and close < ma200) else 0
        daily_fa["_macd_death_cross"] = 1 if (macd_val > 0 and macd_val <= macds_val) else 0

        # 执行
        daily_state = engine.execute(
            portfolio=portfolio,
            decision=decision,
            row=row,
            idx=idx,
            df=df,
            fa_metrics=daily_fa,
        )

        # 收集所有规则的条件详情（用执行引擎注入后的 row_dict）
        eval_dict = row.to_dict()
        eval_dict.update(daily_fa)
        eval_dict["close"] = close
        eval_dict["_close"] = close
        eval_dict["_high"] = float(row.get("high", 0))
        eval_dict["_low"] = float(row.get("low", 0))

        rule_details = {}
        for i, rule in enumerate(rules):
            rid = f"R{i+1}"
            cond_strs = []
            all_ok = True
            for c in rule.conditions:
                val = eval_dict.get(c.field)
                if val is None:
                    ok = False
                    cond_strs.append(f"{c.field}=N/A {c.op.value} {c.value} ❌缺数据")
                else:
                    try:
                        ok = c.evaluate(eval_dict)
                        val_str = f"{float(val):.4f}" if isinstance(val, (int, float)) else str(val)[:20]
                        emoji = "✅" if ok else "❌"
                        cond_strs.append(f"{c.field}={val_str} {c.op.value} {c.value} {emoji}")
                    except Exception as e:
                        ok = False
                        cond_strs.append(f"{c.field}={val} ERR:{e}")
                if not ok:
                    all_ok = False
            rule_details[rid] = {
                "triggered": all_ok,
                "conds": " | ".join(cond_strs),
                "action": rule.action.value,
            }

        triggered_list = [rid for rid, d in rule_details.items() if d["triggered"]]

        # 获取技术指标
        ma50 = float(row_dict.get("close_50_sma", 0))
        ma200 = float(row_dict.get("close_200_sma", 0))
        macd_val = float(row_dict.get("macd", 0))
        macdh_val = float(row_dict.get("macdh", 0))
        macds_val = float(row_dict.get("macds", 0))

        rows.append({
            "日期": date_str,
            "收盘价": round(close, 2),
            "涨跌幅%": round(float(row.get("pct_chg", 0)), 2),
            "50日均线": round(ma50, 2),
            "200日均线": round(ma200, 2),
            "收盘<50日": "⚠️跌破" if close < ma50 else "✅",
            "收盘<200日": "⚠️跌破" if close < ma200 else "✅",
            "MACD": round(macd_val, 4),
            "MACD柱": round(macdh_val, 4),
            "MACD信号": round(macds_val, 4),
            "MACD零轴死叉": "⚠️是" if (macd_val > 0 and macd_val < macds_val) else "—",
            # 每条规则的触发状态
            **{f"{rid}_触发": "🔥是" if rule_details[rid]["triggered"] else "—"
               for rid in [f"R{i+1}" for i in range(len(rules))]},
            "触发规则": ", ".join(triggered_list) if triggered_list else "—",
            "触发数": len(triggered_list),
            # 执行动作
            "动作": daily_state.action,
            "动作价": round(daily_state.action_price, 2) if daily_state.action != "HOLD" else "—",
            "动作量": daily_state.action_shares if daily_state.action != "HOLD" else "—",
            "持仓股": portfolio.shares,
            "现金": round(portfolio.cash, 0),
            "总资产": round(portfolio.cash + portfolio.shares * close, 0),
            # 条件详情（每条规则一个列）
            **{f"{rid}_条件": rule_details[rid]["conds"]
               for rid in [f"R{i+1}" for i in range(len(rules))]},
        })

    result_df = pd.DataFrame(rows)
    print(f"   ✅ 收集到 {len(result_df)} 天的数据")
    triggered_days = result_df[result_df["触发数"] > 0]
    print(f"   触发天数: {len(triggered_days)}")
    for _, td in triggered_days.iterrows():
        print(f"     {td['日期']}: {td['触发规则']} → 动作={td['动作']}")

    # ── 6. 输出 Excel ────────────────────────────────────────────
    print("\n📋 Step 6: 生成 Excel")
    output_path = f"backtest_results/execution_daily_test_{SYMBOL}_{START_DATE}_{END_DATE}.xlsx"
    os.makedirs("backtest_results", exist_ok=True)

    wb = Workbook()
    
    # ── Sheet 1: 每日执行详情 ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "每日执行"

    # 颜色定义
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    triggered_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    triggered_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    action_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    normal_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 写标题行
    base_cols = ["日期", "收盘价", "涨跌幅%", "50日均线", "200日均线",
                 "收盘<50日", "收盘<200日", "MACD", "MACD柱", "MACD信号",
                 "MACD零轴死叉"]
    rule_cols = [f"R{i+1}_触发" for i in range(len(rules))]
    extra_cols = ["触发规则", "触发数", "动作", "动作价", "动作量",
                  "持仓股", "现金", "总资产"]
    
    all_cols = base_cols + rule_cols + extra_cols
    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 写数据
    for row_idx, (_, row_data) in enumerate(result_df.iterrows()):
        for col_idx, col_name in enumerate(all_cols, 1):
            val = row_data.get(col_name, "")
            cell = ws1.cell(row=row_idx + 2, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            
            # 触发的高亮
            if col_name.endswith("_触发") and val == "🔥是":
                cell.fill = triggered_fill
                cell.font = triggered_font
            # 动作高亮
            if col_name == "动作" and val != "HOLD":
                cell.fill = action_fill
                cell.font = Font(name="微软雅黑", size=10, bold=True)

    # 冻结首行
    ws1.freeze_panes = "A2"
    
    # 列宽
    ws1.column_dimensions["A"].width = 12
    for col_idx in range(2, len(all_cols) + 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = 14

    # ── Sheet 2: 规则分析 ──────────────────────────────────────
    ws2 = wb.create_sheet("规则分析")
    
    analysis_cols = ["规则", "动作", "优先级", "条件数", "条件详情", "可评估", "触发天数", "首次触发"]
    for col_idx, col_name in enumerate(analysis_cols, 1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for i, rule in enumerate(rules):
        rid = f"R{i+1}"
        cond_strs = []
        all_evaluable = True
        for c in rule.conditions:
            needs_fa = c.field.startswith("annual_") or c.field.startswith("quarter_") or c.field.startswith("fa_")
            source = "FA" if needs_fa else ("技术" if c.field in ("macd","macdh","macds","rsi") else ("价格" if c.field in ("close","high","low","open","volume","pct_chg") else ("均线" if "sma" in c.field else "❓")))
            cond_strs.append(f"{c.field} {c.op.value} {c.value} [{source}]")

        trigger_col = f"{rid}_触发"
        triggered_count = len(result_df[result_df[trigger_col] == "🔥是"])
        first_trigger = result_df[result_df[trigger_col] == "🔥是"]["日期"].iloc[0] if triggered_count > 0 else "—"

        row_data = [
            f"{rid} {rule.name[:30]}",
            rule.action.value,
            rule.priority,
            len(rule.conditions),
            " AND ".join(cond_strs),
            "✅" if all_evaluable else "⚠️",
            triggered_count,
            first_trigger,
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=i + 2, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["E"].width = 80
    ws2.freeze_panes = "A2"

    # ── Sheet 3: 条件详情（每行一个条件） ────────────────────────
    ws3 = wb.create_sheet("条件详情")
    cond_cols = ["规则", "条件序号", "字段", "操作符", "阈值", "数据来源", "可评估"]
    for col_idx, col_name in enumerate(cond_cols, 1):
        cell = ws3.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    row_num = 2
    for i, rule in enumerate(rules):
        rid = f"R{i+1}"
        for j, c in enumerate(rule.conditions):
            needs_fa = c.field.startswith("annual_") or c.field.startswith("quarter_")
            needs_tech = c.field in ("macd", "macdh", "macds", "rsi", "atr")
            source = "FA指标" if needs_fa else ("技术指标" if needs_tech else "日线数据")
            evaluable = "✅" if (needs_fa or needs_tech or c.field in ("close", "high", "low", "open", "volume", "pct_chg", "turn") or "sma" in c.field) else "⚠️"

            row_data = [f"{rid} {rule.name[:20]}", j + 1, c.field, c.op.value, c.value, source, evaluable]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws3.cell(row=row_num, column=col_idx, value=val)
                cell.font = normal_font
                cell.border = thin_border
            row_num += 1

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["C"].width = 20
    ws3.freeze_panes = "A2"

    wb.save(output_path)
    print(f"   ✅ Excel 已保存: {output_path}")
    print(f"   Sheets: 每日执行 | 规则分析 | 条件详情")
    
    return output_path


if __name__ == "__main__":
    output = main()
    print(f"\n🎉 完成！输出文件: {output}")
